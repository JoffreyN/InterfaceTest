from pyDes import *
from dubboTelnet import Dubbo
from email.header import Header
from openpyxl import load_workbook
from urllib.parse import urlencode
from email.mime.text import MIMEText
import os,sys,logging,re,time,random,json,http.client,hashlib,base64,smtplib,faker,requests
fake=faker.Faker(locale='zh_CN')

def getCellStr(cell,noStrip=0):
	if cell.value==None:return cell.value
	try:
		strs=str(cell.value) if noStrip else str(cell.value).strip()
	except AttributeError:
		strs=str(cell.value)
	return strs

#获取并执行测试用例
def runTest(testCaseFile,startTime):
	start=time.perf_counter()#开始计时
	checkModeDic={'Contain':'包含','Equal':'相等'}
	resultJson={"testName":"","beginTime":time.strftime('%Y-%m-%d %H:%M:%S',startTime),"testAll":0,"testPass":0,"testFail":0,"testSkip":0,"totalTime":"","testResult":[]}
	resultJson["testName"]=os.path.split(testCaseFile)[-1].split('.')[0]#用例名称
	testCaseFile=os.path.join(os.getcwd(),testCaseFile)
	if not os.path.exists(testCaseFile):
		logging.error(f'{testCaseFile}用例文件不存在！！！\n')
		sys.exit()
	excel=load_workbook(testCaseFile)
	table=excel.active
	resultJson["testAll"]=table.max_row-1#总用例数
	correlationDict={}
	correlationDict['${hashPassword}']=hash1Encode('123456')
	correlationDict['${session}']=None
	for row in table.iter_rows(min_row=2):
		correlationDict['${randomEmail}']=fake.email()
		correlationDict['${randomTel}']=fake.phone_number()
		correlationDict['${timestamp}']=int(time.time())
		num=getCellStr(row[0])
		interfaceName=getCellStr(row[1])
		interFaceType=getCellStr(row[2])
		host=getCellStr(row[3])
		reqPath=getCellStr(row[4])
		reqType=getCellStr(row[5])
		reqHeader=getCellStr(row[6])
		reqDataType=getCellStr(row[7])
		reqData=getCellStr(row[8])
		encryption=getCellStr(row[9])
		checkMode=getCellStr(row[10])
		checkPoint=getCellStr(row[11],1)
		correlation=getCellStr(row[12])
		active=getCellStr(row[13])
		if active!='Yes':
			resultJson["testSkip"]+=1
			resultJson["testResult"].append({
				"className":interfaceName,
				"methodName":interFaceType,
				"description":f"{interFaceType.lower()}://{host}/{reqPath}"[::-1].replace('//','/',1)[::-1],"spendTime":"0 ms","status":"跳过","log":["跳过"]
				})
			continue
		for key in correlationDict:
			if reqPath.find(key)>0:
				reqPath=reqPath.replace(key,str(correlationDict[key]))
		if interFaceType.upper() in ['HTTP','HTTPS']:
			if reqDataType=='Form':
				if reqData:#reqData不为空时
					dataFile=reqData
					if os.path.exists(dataFile):
						fopen=open(dataFile,encoding='utf-8')
						reqData=fopen.readline()
						fopen.close()
					for keyword in correlationDict:
						if reqData.find(keyword)>0:
							reqData=reqData.replace(keyword,str(correlationDict[keyword]))
					try:
						if encryption=='MD5':
							reqData=json.loads(reqData)
							status,md5=getMD5(host,urlencode(reqData).replace("%27","%22"))
							if status!=200:
								# logging.error(num+' '+interfaceName+"[ "+str(status)+" ], 获取md5验证码失败！！！\n")
								logging.error(f"{num} {interfaceName}[{str(status)}], 获取md5验证码失败！！！\n")
								continue
							reqData=dict(reqData,**{"sign":md5.decode("utf-8")})
							reqData=urlencode(reqData).replace("%27","%22")
						elif encryption=='DES':
							reqData=json.loads(reqData)
							reqData=urlencode({'param':encodePostStr(reqData)})
						else:
							reqData=urlencode(json.loads(reqData))
					except Exception as e:
						logging.error(f'{num} {interfaceName} 请求的数据有误，请检查[ReqData]字段是否是标准的json格式字符串！\n')
						continue
			elif reqDataType=='Data':
				dataFile=reqData
				if os.path.exists(dataFile):
					fopen=open(dataFile,encoding='utf-8')
					reqData=fopen.readline()
					fopen.close()
				for keyword in correlationDict:
					if reqData.find(keyword)>0:
						reqData=reqData.replace(keyword,str(correlationDict[keyword]))
				reqData=reqData.encode('utf-8')
			elif reqDataType=='File':
				dataFile=reqData
				if not os.path.exists(dataFile):
					logging.error(f'{num} {interfaceName} 文件路径配置无效，请检查[Request Data]字段配置的文件路径是否存在！！！\n')
					continue
				fopen=open(dataFile,'rb')
				data=fopen.read()
				fopen.close()
				reqData='''
------WebKitFormBoundaryDf9uRfwb8uzv1eNe
Content-Disposition:form-data;name="file";filename="%s"
Content-Type:
Content-Transfer-Encoding:binary

%s
------WebKitFormBoundaryDf9uRfwb8uzv1eNe--
''' % (os.path.basename(dataFile),data)
			_start=time.perf_counter()
			status,resp=interfaceTest(num,interfaceName,interFaceType,host,reqPath,reqData,checkMode,checkPoint,reqType,reqHeader,reqDataType,correlationDict['${session}'])
			_spendTime=f"{int((time.perf_counter()-_start)*1000)} ms"
			if status==200:#成功
				resultJson["testPass"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}{reqPath}",
					"spendTime":_spendTime,
					"status":"成功",
					"log": [f"请求数据：{reqData}",f"预期结果：{checkPoint}",f"检查方式：{checkModeDic[checkMode]}",f"返回数据：{resp}"]
					})
			elif status==2001:#返回数据与预期不一致
				resultJson["testFail"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}{reqPath}",
					"spendTime":_spendTime,
					"status":"失败",
					"log": [f"请求数据：{reqData}",f"预期结果：{checkPoint}",f"检查方式：{checkModeDic[checkMode]}",f"返回数据：{resp}"]
					})
				continue
			elif status==0:#连接错误
				resultJson["testFail"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}{reqPath}",
					"spendTime":_spendTime,
					"status":"失败",
					"log": [f"接口连接异常：{resp}"]
					})
				continue
			elif status==400:#数据错误
				resultJson["testFail"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}{reqPath}",
					"spendTime":_spendTime,
					"status":"失败",
					"log": [f"数据错误：{resp}"]
					})
				continue
			else:
				resultJson["testFail"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}{reqPath}",
					"spendTime":_spendTime,
					"status":"失败",
					"log": [f"接口返回异常：{resp}"]
					})
				continue
			if correlation:
				for j in range(len(correlation)):
					param=correlation[j].split('=')
					if len(param)==2:
						if param[1]=='' or not re.search(r'^\[',param[1]) or not re.search(r'\]$',param[1]):
							logging.error(f'{num} {interfaceName} 关联参数设置有误，请检查[Correlation]字段参数格式是否正确！！！\n')
							continue
						value=resp
						for key in param[1][1:-1].split(']['):
							try:
								temp=value[int(key)]
							except:
								try:temp=value[key]
								except:break
							value=temp
						correlationDict[param[0]]=value
		elif interFaceType=='Dubbo':
			if reqType=='Telnet':
				try:
					host,portStr=host.split(':')
				except ValueError:
					print('Error:',interfaceName,'请检查Host/IP_Port字段，格式为IP:Port')
					resultJson["testFail"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":"0 ms",
						"status":"失败",
						"log": [f"数据错误，请检查Host/IP_Port字段，格式应为IP:Port：{host}"]
						})
					logging.error(f'{num} {interfaceName} Host/IP_Port字段格式错误\n')
					continue
				try:
					service,method=reqPath.split('#')
				except ValueError:
					print('Error:',interfaceName,'请检查ReqPath|Service#method字段，格式为 接口名#方法')
					resultJson["testFail"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":"0 ms",
						"status":"失败",
						"log": [f"数据错误，请检查ReqPath|Service#method字段，格式应为 接口名#方法：{interfaceName}"]
						})                    
					logging.error(f'{num} {interfaceName} ReqPath|Service#method字段格式错误\n')
					continue
				_start=time.perf_counter()
				status,resp=interfaceDubbloTelnet(num,interfaceName,host,int(portStr),service,method,reqData,checkPoint,checkMode)
				_spendTime=f"{int((time.perf_counter()-_start)*1000)} ms"
				if status==200:
					resultJson["testPass"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":_spendTime,
						"status":"成功",
						"log": [f"请求数据：{reqData}",f"预期结果：{checkPoint}",f"检查方式：{checkModeDic[checkMode]}",f"返回数据：{resp}"]
						})
				elif status==2001:#返回数据与预期不一致
					resultJson["testFail"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":_spendTime,
						"status":"失败",
						"log": [f"请求数据：{reqData}",f"预期结果：{checkPoint}",f"检查方式：{checkModeDic[checkMode]}",f"返回数据：{resp}"]
						})
				elif status==500:#无法连接目标计算机！
					resultJson["testFail"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":_spendTime,
						"status":"失败",
						"log": ["失败原因：无法连接目标计算机！"]
						})
				elif status==503:#发送数据异常
					resultJson["testFail"]+=1
					resultJson["testResult"].append({
						"className":interfaceName,
						"methodName":interFaceType,
						"description":f"{interFaceType.lower()}://{host}:{portStr}/{reqPath}",
						"spendTime":_spendTime,
						"status":"失败",
						"log": [f"发送数据异常：{e}"]
						})
			elif reqType=='Hessian':
				resultJson["testFail"]+=1
				resultJson["testResult"].append({
					"className":interfaceName,
					"methodName":interFaceType,
					"description":f"{interFaceType.lower()}://{host}/{reqPath}",
					"spendTime":'0 ms',
					"status":"失败",
					"log": ["失败原因：Hessian方法暂未开发，敬请期待！"]
					})
	resultJson["totalTime"]=f'{int((time.perf_counter()-start)*1000)} ms'
	# resultJson["totalTime"]=f'{int(time.perf_counter()-start)} s'
	return resultJson

def interfaceDubbloTelnet(num,interfaceName,host,port,service,method,reqData,checkPoint,checkMode):#dubbo接口telnet方式测试
	try:
		conn=Dubbo(host,port)
	except (ConnectionRefusedError,TimeoutError):
		logging.error(f'{num} {interfaceName} 无法连接目标计算机！\n')
		return 500,'无法连接目标计算机！'
	try:
		result=conn.invoke(service,method,reqData)
	except Exception as e:
		logging.error(f'{num} {interfaceName} 发送数据异常!\n')
		return 503,e
	finally:
		conn.close()
	status,response=check(num,interfaceName,result,checkPoint,checkMode)
	return status,response

# HTTP、HTTPS接口测试
def interfaceTest(num,interfaceName,interFaceType,host,reqPath,reqData,checkMode,checkPoint,reqType,reqHeader,reqDataType,session):
	head={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36','Content-Type':'application/x-www-form-urlencoded; charset=UTF-8'}
	if reqHeader:
		try:
			head=dict(head,**eval(reqHeader))
		except TypeError:
			return 400,'请检查ReqHeader字段格式是否为字典'
	if session is not None:
		head['Cookie']=f'session={session}'
		if reqDataType=='File':
			head['Content-Type']='multipart/form-data;boundary=----WebKitFormBoundaryDf9uRfwb8uzv1eNe;charset=UTF-8'
		elif reqDataType=='Data':
			head['Content-Type']='text/plain; charset=UTF-8'
	if reqType=='POST':
		try:
			req=requests.post(f'{interFaceType.lower()}://{host}{reqPath}',data=reqData,headers=head,timeout=10)
		except Exception as e:
			logging.error(f'{num} {interfaceName} 连接异常\n')
			return 0,e
	elif reqType=='GET':
		try:
			if reqData:
				req=requests.get(f'{interFaceType.lower()}://{host}{reqPath}?{reqData}',headers=head,timeout=10)
			else:
				req=requests.get(f'{interFaceType.lower()}://{host}{reqPath}',headers=head,timeout=10)
		except Exception as e:
			logging.error(f'{num} {interfaceName} 连接异常\n')
			return 0,e
	status=req.status_code
	resp=req.text
	if status==200:
		status,resp=check(num,interfaceName,resp,checkPoint,checkMode)
	else:
		logging.error(f'{num} {interfaceName} 失败！ [{str(status)}], {resp}\n')
	return status,resp

def check(num,interfaceName,response,checkPoint,checkMode):#检查返回的结果
	if checkMode=='Equal':
		flag=True if str(response)==str(checkPoint) else False
	elif checkMode=='Contain':
		flag=True if str(checkPoint) in str(response) else False
	if flag:
		logging.info(f'{num} {interfaceName} 成功, {response}\n')
		return 200,response
	else:
		logging.error(f'{num} {interfaceName} 失败！response: {response}\n')
		return 2001,response

#获取md5验证码
def getMD5(url,postData):
	head={'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8','X-Requested-With':'XMLHttpRequest'}
	conn=http.client.HTTPConnection('this.ismyhost.com')
	conn.request('POST','/get_isignature',postData,headers=head)
	response=conn.getresponse()
	return response.status,response.read()

# hash1加密
def hash1Encode(codeStr):
	hashobj=hashlib.sha1()
	hashobj.update(codeStr.encode('utf-8'))
	return hashobj.hexdigest()

# DES加密
def desEncode(desStr):
	k=des('secretKEY',padmode=PAD_PKCS5)
	encodeStr=base64.b64encode(k.encrypt(json.dumps(desStr)))
	return encodeStr

# 字典排序
def encodePostStr(postData):
	keyDict={'key':'secretKEY'}
	mergeDict=dict(postData,**keyDict)
	mergeDict=sorted(mergeDict.items())
	postStr=''
	for i in mergeDict:
		# postStr=postStr+i[0]+'='+i[1]+'&'
		postStr=f'{postStr}{i[0]}={i[1]}&'
	postStr=postStr[:-1]
	hashobj=hashlib.sha1()
	hashobj.update(postStr.encode('utf-8'))
	token=hashobj.hexdigest()
	postData['token']=token
	return desEncode(postData)

#发送通知邮件
def sendMail(text):
	sender='1101170736@qq.com'
	receiver=['zoupeng-jc@bestpay.com.cn']
	mailToCc=['2806646694@qq.com']
	# mailToCc=['yuhao@bestpay.com.cn','2806646694@qq.com']
	subject='[AutoInterfaceTest]接口自动化测试报告通知'
	# smtpserver='smtp.exmail.qq.com'
	# smtpserver='smtp.qq.com'
	username='1101170736@qq.com'
	password='qwibtbluzipljihc'
	# password='Zp19940130xyz'
	
	msg=MIMEText(text,'html','utf-8')
	# msg['Subject']=subject
	msg['Subject']=Header(subject,'utf-8') #设置主题和格式
	# msg['From']=sender
	msg['From']=Header(sender,'utf-8') #设置显示在邮件里的发件人
	# msg['To']=';'.join(receiver)
	msg['To']=Header(';'.join(receiver),'utf-8') #设置显示在邮件里的收件人
	# msg['Cc']=';'.join(mailToCc)
	msg['Cc']=Header(';'.join(mailToCc),'utf-8')
	# smtp=smtplib.SMTP()
	smtp=smtplib.SMTP_SSL('smtp.qq.com',465)
	# smtp.connect(smtpserver)
	smtp.login(username, password)
	smtp.sendmail(sender, receiver, msg.as_string())
	smtp.quit()

# def out(strings):
# 	sys.stdout.write(f'{strings}\r')
# 	sys.stdout.flush()

def main(testCaseFile):
	startTime=time.localtime(time.time())#开始测试时间
	timeStamp=time.strftime('%Y%m%d%H%M%S',startTime)#时间，用于命名保存的文件
	resultJson=runTest(f'TestCase/{testCaseFile}',startTime)
	with open(f"TestReport/data/data_{timeStamp}.js",'w',encoding='utf-8') as f:#保存结果json
		f.write(f'var resultData = {resultJson};')
	html=open('TestReport/reportDemo.html','r',encoding='utf-8').read().replace('dataDemo',f"data_{timeStamp}")
	htmlPath=f"TestReport/{testCaseFile.split('.')[0]}_report_{timeStamp}.html"
	with open(htmlPath,'w',encoding='utf-8') as f:
		f.write(html)
	# out('正在自动发送测试报告……')
	# sendMail(html)
	return htmlPath

if __name__=='__main__':
	if len(sys.argv)==1:#传入TestCase目录下的测试用例名称
		testCaseFileList=input('输入TestCase目录下的测试用例名称，如有多个空格分隔：').strip().split()
		testCaseFileList=testCaseFileList if testCaseFileList else ['demo.xlsx']
	elif len(sys.argv)==2:#传入TestCase目录下的测试用例名称
		testCaseFileList=sys.argv[1].strip().split() if sys.argv[1].strip().endswith('xlsx') else ['demo.xlsx']
	else:
		print('输入有误！')
		sys.exit()
	for testCaseFile in testCaseFileList:
		# 1、创建一个logger
		logger=logging.getLogger('')
		logger.setLevel(logging.DEBUG)
		# 2、定义handler的输出格式（formatter
		formatter=logging.Formatter('[%(asctime)s] [%(levelname)s]: %(message)s')
		# 3、创建一个handler，用于写入日志文件
		fh=logging.FileHandler(filename=os.path.join(os.getcwd(),f'log/{os.path.splitext(testCaseFile)[0]}.log'),mode='a+',encoding='utf-8')
		fh.setLevel(logging.DEBUG)
		# 4、创建一个handler，用于输出到控制台
		ch=logging.StreamHandler()
		ch.setLevel(logging.DEBUG)
		# 5、给handler添加formatter
		fh.setFormatter(formatter)
		ch.setFormatter(formatter)
		# 6、给logger添加handler
		logger.addHandler(fh)
		logger.addHandler(ch)
		reportPath=main(testCaseFile)
		print(f"用例 {testCaseFile} 执行完成！ {time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))}")
		# os.system(f'start file:///{os.getcwd()}/{reportPath}')